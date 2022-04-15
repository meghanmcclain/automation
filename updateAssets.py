#! python3
# updateAssets.py - Corrects Cost Center & Fund Names in All Assets spreadsheet
# downloaded the All Assets.xlsx spreadsheet from Teams SAP Data Dump (Joe's message)
#go through 14,792 rows and update the Cost Centers and Funds with their proper names
#shouldn't use find-and-replace - that will update everywhere in the worksheet that it sees the value, not just a specific column
#just in case, for confidentiality reasons, I replaced the cost center and fund numbers with 10-digit numbers that I made up

import openpyxl, threading, time, os, os.path
from openpyxl.utils import column_index_from_string

print('Start of program.')

wb = openpyxl.load_workbook('allAssetsOriginal.xlsx')
sheet = wb['Sheet1']

# The Cost Center numbers and their proper names
#most of the values in the original spreadsheet are strings but two of them are integers 
COST_CENTER_UPDATES = {'0000000001': 'MAS',
                '0000000002': 'DNS',
                '0000000003': 'Reserve',
                '0000000004': 'MAS',
                '0000000005': 'MAS',
                '0000000006': 'MAS',
                '0000000007': 'MAS',
                '0000000008': 'MAS',
                '0000000009': 'NFS',
                '0000000010': 'Ops',
                '0000000011': 'Reserve',
                '0000000012': 'Ops',
                '0000000013': 'SWIC',
                '0000000014': 'Reserve',
                '0000000015': 'DNS',
                '0000000016': 'DNS',
                '0000000017': 'DNS',
                '0000000018': 'DNS',
                '0000000019': 'DNS',
                '0000000020': 'DNS',
                '0000000021': 'Ops',
                '0000000022': 'NFS',
                '0000000023': 'NFS',
                '0000000024': 'NFS',
                '0000000025': 'NFS',
                '0000000026': 'NFS',
                '0000000027': 'MAS',
                '0000000028': 'MAS',
                '0000000029': 'PGA',
                '0000000030': 'PGA',
                '0000000031': 'PGA',
                '0000000032': 'DNS',
                '0000000033': 'Ops',
                0000000008: 'MAS',
                0000000015: 'DNS'}

FUND_UPDATES = {'0000000034': 'General Revenue Fund',
                '0000000035': 'Radiation Protection Fund',
                '0000000036': 'Emergency Planning and Training Fund',
                '0000000037': 'Indoor Radon Mitigation Fund',
                '0000000038': 'State Coronavirus Urgent Remediation Emergency Fund',
                '0000000039': 'Nuclear Civil Protection Planning Fund',
                '0000000040': 'Federal Aid Disaster Fund',
                '0000000041': 'Federal Civil Preparedness Administrative Fund',
                '0000000042': 'September 11th Fund',
                '0000000043': 'Disaster Response and Recovery Fund',
                '0000000044': 'Homeland Security Emergency Preparedness Trust Fund',
                '0000000045': 'Homeland Security Emergency Preparedness Trust Fund',
                '0000000046': 'Nuclear Safety Emergency Preparedness Fund',
                '0000000047': 'Sheffield February 1982 Agreed Order Fund',
                '0000000048': 'Low-Level Radioactive Waste Facility Development and Operation Fund'}

def updateFunds(): 
    #Loop through the rows and update the Fund Column
    for rowNum in range(2, sheet.max_row): #skip the first row
        fundNumber = sheet.cell(row=rowNum, column=column_index_from_string('P')).value
        if fundNumber in FUND_UPDATES:
            sheet.cell(row=rowNum, column=column_index_from_string('P')).value = FUND_UPDATES[fundNumber]
    print('Fund Update - Complete')

#Create and start the Thread objects
threadObjects = [] #a list of all the Thread objects
threadObj = threading.Thread(target=updateFunds) #don't call updateFunds() yet
threadObjects.append(threadObj)
threadObj.start() #now call updateFunds(), second thread starts

#Loop through the rows and update the Cost Center Column
for rowNum in range(2, sheet.max_row): #skip the first row
    costCenterNumber = sheet.cell(row=rowNum, column=column_index_from_string('V')).value
    if costCenterNumber in COST_CENTER_UPDATES:
        sheet.cell(row=rowNum, column=column_index_from_string('V')).value = COST_CENTER_UPDATES[costCenterNumber]
print('Cost Center Update - Complete')

#Wait for all threads to end before saving to a new file
for threadObj in threadObjects:
    threadObj.join()
print('Both threads are Done.')

newFileName = 'updatedAllAssets.xlsx'

print('File can be found here: '+ os.getcwd() + '\\' + newFileName)

fileExists = os.path.exists(newFileName)

#check to see if file already exists, if it does, overwrite it
if fileExists == True:
    os.remove(newFileName)

wb.save(newFileName)
